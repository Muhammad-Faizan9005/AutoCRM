from __future__ import annotations

import csv
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

from fastapi import HTTPException, status
from supabase import Client

from app.repositories.customer_repository import CustomerRepository
from app.repositories.lead_repository import LeadRepository
from app.repositories.note_repository import NoteRepository
from app.repositories.organization_repository import OrganizationRepository
from app.repositories.ticket_repository import TicketRepository
from app.schemas.lead import LeadCreate
from app.schemas.imports import ImportFailure, ImportResult
from app.schemas.ticket import TicketCreate
from app.utils.team_access import get_agent_team_id


try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - validated during runtime if dependency is missing
    load_workbook = None


SUPPORTED_EXTENSIONS = {".csv", ".xlsx", ".xlsm"}
NOTE_HEADER_TOKENS = {"note", "notes", "comment", "comments", "remark", "remarks"}


@dataclass
class ParsedRow:
    row_number: int
    values: dict[str, Any]


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    normalized = str(value).strip().lower().replace(" ", "_").replace("-", "_")
    return normalized


def _normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_note_header(header: str) -> bool:
    for segment in header.split("_"):
        cleaned_segment = segment.rstrip("0123456789")
        if cleaned_segment in NOTE_HEADER_TOKENS:
            return True
    return False


def _extract_note_content(values: dict[str, Any]) -> str | None:
    note_fragments: list[str] = []
    for header, value in values.items():
        if not _is_note_header(header):
            continue

        fragment = _normalize_cell(value)
        if fragment:
            note_fragments.append(fragment)

    if not note_fragments:
        return None

    return " | ".join(note_fragments)


def _parse_csv(file_bytes: bytes) -> list[ParsedRow]:
    last_error: Exception | None = None
    text: str | None = None

    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = file_bytes.decode(encoding)
            break
        except UnicodeDecodeError as exc:
            last_error = exc

    if text is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unable to decode CSV file") from last_error

    reader = csv.DictReader(StringIO(text))
    if not reader.fieldnames:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="CSV file is missing header row")

    normalized_headers = [_normalize_header(h) for h in reader.fieldnames]
    if any(not h for h in normalized_headers):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="CSV header contains empty column name")

    parsed_rows: list[ParsedRow] = []
    for row_index, row in enumerate(reader, start=2):
        normalized_row = {}
        for original_key, normalized_key in zip(reader.fieldnames, normalized_headers):
            normalized_row[normalized_key] = _normalize_cell(row.get(original_key))

        if all(not value for value in normalized_row.values()):
            continue

        parsed_rows.append(ParsedRow(row_number=row_index, values=normalized_row))

    return parsed_rows


def _parse_excel(file_bytes: bytes) -> list[ParsedRow]:
    if load_workbook is None:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel import support is unavailable. Install openpyxl.",
        )

    workbook = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    header_row = next(iterator, None)

    if not header_row:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel file is empty")

    normalized_headers = [_normalize_header(value) for value in header_row]
    if not any(normalized_headers):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel header row is empty")
    if any(not h for h in normalized_headers):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel header contains empty column name")

    parsed_rows: list[ParsedRow] = []
    for row_index, values in enumerate(iterator, start=2):
        row_dict: dict[str, str] = {}
        for column_index, header in enumerate(normalized_headers):
            cell_value = values[column_index] if column_index < len(values) else None
            row_dict[header] = _normalize_cell(cell_value)

        if all(not value for value in row_dict.values()):
            continue

        parsed_rows.append(ParsedRow(row_number=row_index, values=row_dict))

    return parsed_rows


def _parse_rows(file_name: str, file_bytes: bytes) -> list[ParsedRow]:
    suffix = Path(file_name or "").suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file type. Use CSV or Excel (.xlsx/.xlsm).",
        )

    if suffix == ".csv":
        return _parse_csv(file_bytes)
    return _parse_excel(file_bytes)


class ImportService:
    def __init__(self, db: Client):
        self.customer_repository = CustomerRepository(db)
        self.lead_repository = LeadRepository(db)
        self.note_repository = NoteRepository(db)
        self.organization_repository = OrganizationRepository(db)
        self.ticket_repository = TicketRepository(db)

    async def get_or_create_organization(
        self,
        *,
        company_name: str,
        owner_id: str | None = None,
    ) -> dict[str, Any] | None:
        """
        Find or create an organization by company name.
        Returns the organization dict or None if company_name is empty.
        """
        if not company_name or not company_name.strip():
            return None

        company_name_clean = company_name.strip()
        existing_org = await self.organization_repository.find_one(filters={"name": company_name_clean})
        if existing_org:
            return existing_org

        org_data = {"name": company_name_clean}
        if owner_id:
            org_data["owner_id"] = owner_id
            org_data["team_id"] = await get_agent_team_id(self.organization_repository.db, owner_id)
        created_org = await self.organization_repository.create(org_data)
        return created_org

    async def create_or_update_lead_from_row(
        self,
        *,
        row: dict[str, Any],
        notes_content: str | None,
        owner_id: str | None = None,
    ) -> tuple[dict[str, Any] | None, bool]:
        """
        Create or update a lead from imported row data.
        Returns the saved lead dict and a flag indicating whether it was created.
        """
        organization_id = None
        company = row.get("company")
        if company:
            org = await self.get_or_create_organization(company_name=company, owner_id=owner_id)
            if org:
                organization_id = org.get("id")

        lead_data = {
            "name": row.get("name", ""),
            "email": row.get("email") or None,
            "phone": row.get("phone") or None,
            "company": row.get("company") or None,
            "source": row.get("source") or "import",
            "status": row.get("status") or "new",
            "organization_id": str(organization_id) if organization_id else None,
            "score": row.get("score"),
            "score_reason": row.get("score_reason") or None,
        }
        if owner_id:
            lead_data["owner_id"] = owner_id

        model = LeadCreate(**lead_data)
        payload = model.model_dump(exclude_none=True)

        saved_lead: dict[str, Any]
        created = False
        if model.email:
            existing_lead = await self.lead_repository.find_one(filters={"email": str(model.email)})
            if existing_lead:
                if existing_lead.get("owner_id"):
                    payload.pop("owner_id", None)
                saved_lead = await self.lead_repository.update_by_id(existing_lead["id"], payload)
            else:
                saved_lead = await self.lead_repository.create(payload)
                created = True
        else:
            saved_lead = await self.lead_repository.create(payload)
            created = True

        if notes_content:
            saved_lead = dict(saved_lead)
            saved_lead["notes"] = notes_content

        return saved_lead, created

    async def create_notes_from_lead(self, *, lead: dict[str, Any]) -> dict[str, Any] | None:
        """
        Create a note from lead notes field.
        Returns the created note dict or None if lead has no notes.
        """
        lead_id = lead.get("id")
        notes_content = lead.get("notes")

        if not lead_id or not notes_content or not notes_content.strip():
            return None

        note_content = notes_content.strip()
        existing_note = await self.note_repository.find_one(
            filters={"entity_type": "lead", "entity_id": str(lead_id), "content": note_content}
        )
        if existing_note:
            return existing_note

        note_data = {
            "entity_type": "lead",
            "entity_id": str(lead_id),
            "content": note_content,
            "author_id": None,
        }

        created_note = await self.note_repository.create(note_data)
        return created_note

    async def ingest_lead_payload(self, *, payload: dict[str, Any], owner_id: str | None = None) -> dict[str, Any]:
        notes_content = _extract_note_content(payload)
        lead_payload = {
            "name": (
                payload.get("name")
                or payload.get("full_name")
                or payload.get("lead_name")
                or payload.get("company")
                or (payload.get("email", "").split("@")[0] if payload.get("email") else "")
            ),
            "email": payload.get("email") or None,
            "phone": payload.get("phone") or None,
            "company": payload.get("company") or None,
            "source": payload.get("source") or "import",
            "status": payload.get("status") or "new",
            "score": payload.get("score") or None,
            "score_reason": payload.get("score_reason") or None,
        }

        lead, _created = await self.create_or_update_lead_from_row(
            row=lead_payload,
            notes_content=notes_content,
            owner_id=owner_id,
        )
        if lead:
            await self.create_notes_from_lead(lead={**lead, "notes": notes_content} if notes_content else lead)

        return lead

    async def import_leads(self, *, file_name: str, file_bytes: bytes, owner_id: str | None = None) -> ImportResult:
        parsed_rows = _parse_rows(file_name=file_name, file_bytes=file_bytes)

        created_count = 0
        updated_count = 0
        failures: list[ImportFailure] = []

        for parsed_row in parsed_rows:
            try:
                notes_content = _extract_note_content(parsed_row.values)
                payload = {
                    "name": (
                        parsed_row.values.get("name")
                        or parsed_row.values.get("full_name")
                        or parsed_row.values.get("lead_name")
                        or parsed_row.values.get("company")
                        or (parsed_row.values.get("email", "").split("@")[0] if parsed_row.values.get("email") else "")
                    ),
                    "email": parsed_row.values.get("email") or None,
                    "phone": parsed_row.values.get("phone") or None,
                    "company": parsed_row.values.get("company") or None,
                    "source": parsed_row.values.get("source") or "import",
                    "status": parsed_row.values.get("status") or "new",
                    "score": parsed_row.values.get("score") or None,
                    "score_reason": parsed_row.values.get("score_reason") or None,
                }
                lead_to_process, was_created = await self.create_or_update_lead_from_row(
                    row=payload,
                    notes_content=notes_content,
                    owner_id=owner_id,
                )
                if lead_to_process and lead_to_process.get("id"):
                    if was_created:
                        created_count += 1
                    else:
                        updated_count += 1

                try:
                    await self.create_notes_from_lead(lead=lead_to_process)
                except Exception:
                    pass

            except Exception as exc:
                failures.append(ImportFailure(row_number=parsed_row.row_number, reason=str(exc)))

        successful_rows = created_count + updated_count
        return ImportResult(
            entity="leads",
            file_name=file_name,
            total_rows=len(parsed_rows),
            successful_rows=successful_rows,
            created_count=created_count,
            updated_count=updated_count,
            failed_count=len(failures),
            failures=failures,
        )

    async def import_tickets(self, *, file_name: str, file_bytes: bytes) -> ImportResult:
        parsed_rows = _parse_rows(file_name=file_name, file_bytes=file_bytes)

        created_count = 0
        failures: list[ImportFailure] = []

        for parsed_row in parsed_rows:
            try:
                customer_id = parsed_row.values.get("customer_id") or None
                customer_email = parsed_row.values.get("customer_email") or None

                if not customer_id and customer_email:
                    customer = await self.customer_repository.find_one(filters={"email": customer_email})
                    if not customer:
                        raise ValueError(f"Customer not found for email: {customer_email}")
                    customer_id = customer["id"]

                if not customer_id:
                    raise ValueError("Either customer_id or customer_email is required")

                payload = {
                    "customer_id": str(customer_id),
                    "subject": parsed_row.values.get("subject", ""),
                    "description": parsed_row.values.get("description") or None,
                    "status": (parsed_row.values.get("status") or "open").lower(),
                    "priority": (parsed_row.values.get("priority") or "medium").lower(),
                    "category": parsed_row.values.get("category") or None,
                }
                model = TicketCreate(**payload)
                create_payload = model.model_dump()
                create_payload["customer_id"] = str(create_payload["customer_id"])

                created_ticket = await self.ticket_repository.create(create_payload)
                assigned_to = parsed_row.values.get("assigned_to") or None
                if assigned_to:
                    await self.ticket_repository.update_by_id(created_ticket["id"], {"assigned_to": assigned_to})

                created_count += 1
            except Exception as exc:
                failures.append(ImportFailure(row_number=parsed_row.row_number, reason=str(exc)))

        return ImportResult(
            entity="tickets",
            file_name=file_name,
            total_rows=len(parsed_rows),
            successful_rows=created_count,
            created_count=created_count,
            updated_count=0,
            failed_count=len(failures),
            failures=failures,
        )
        
